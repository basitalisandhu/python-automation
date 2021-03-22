import os
import time
import schedule
import openpyxl
from mysql.connector import connect
import datetime


def compare():
    book = openpyxl.load_workbook('TableName/tablename.xlsx')

    sheet = book.active
    max_rows = sheet. max_row
    i = 0
    rows = []
    data = []
    row_data = []
    # if not os.path.exists('Details'):
    #     os.makedirs('Details')
    # if not os.path.exists('Count'):
    #     os.makedirs('Count')
    # if not os.path.exists('Table_Name'):
    #     os.makedirs('TableName')
    j=0
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=max_rows, max_col=1):
        i = 0
        print(j)
        j+=1
        for cell in row:
            
            print(cell.value, end=" ")
            print()  

            # dataDict[key] = value

            keyword =cell.value # ask the user for keyword, use raw_input() on Python 2.x
            
            root_dir = "/home/basit/lampstack/apache2/htdocs/fwt"  # path to the root directory to search
            for root, dirs, files in os.walk(root_dir, onerror=None):  # walk the root dir
                for filename in files:  # iterate over the files in the current dir

                    if filename.endswith('.php'):
                        file_path = os.path.join(root, filename)  # build the file path
                        try:
                            # print(file_path)
                            # print(filename)
                            
                            with open(file_path, "rb") as f:  # open the file for reading
                                # read the file line by line
                                for line in f:  # use: for i, line in enumerate(f) if you need line numbers
                                    try:
                                        # print(line)
                                        line = line.decode("utf-8")  # try to decode the contents to utf-8
                                    except ValueError:  # decoding failed, skip the line
                                        continue
                                    if keyword in line:  # if the keyword exists on the current line...
                                        i+=1
                                        # rowss = repr(keyword),repr(line),repr(file_path)  # print the file path
                                        # rows.append(rowss)
                                        # PATH = 'Details/details.xlsx'
                                        # if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
                                        #     wbb = openpyxl.load_workbook(PATH)
                                        #     sheetdetails = wbb["Sheet"]

                                        #     sheetdetails.append(rowss)
                                        # else:
                                        #     wbb = openpyxl.Workbook()
                                        #     wss = wbb.active
                                        #     column_names = ["Table Names","line","path"]
                                        #     wss.append(column_names)
                                        #     sheetdetails = wbb["Sheet"]

                                        #     sheetdetails.append(rowss)

                                        # # data = []
                                        # wbb.save(PATH)
                                        
                        except (IOError, OSError):  # ignore read and permission errors
                            continue
        row = keyword,i  # print the file path
        print(row)
        data.append(row)
        column_names = ["Table Names","count"]

        PATH = 'Count/counts.xlsx'
        if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
            wb = openpyxl.load_workbook(PATH)
            sheet = wb["Sheet"]

            for row in data:
                sheet.append(row)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(column_names)

            sheet = wb["Sheet"]

            for row in data:
                sheet.append(row)

        data = []
        wb.save(PATH)

compare()

# def gen():
#     data = compare()
#     column_names = ["Table Names","count"]

#     PATH = './log_file.xlsx'
#     if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
#         wb = openpyxl.load_workbook('log_file.xlsx')
#         sheet = wb["Sheet"]

#         for row in data:
#             sheet.append(row)
#     else:
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.append(column_names)

#         sheet = wb["Sheet"]

#         for row in data:
#             sheet.append(row)
    
#     wb.save('log_file.xlsx')



# gen()
