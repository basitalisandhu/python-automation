import time
import schedule
import openpyxl
import os
from mysql.connector import connect


conn = connect(host="127.0.0.1",user="root",passwd="12345678")
cursor = conn.cursor()
def create_row():
    newtablename = ('backup_db.dollar_2000')
    oldtablename = ('logi.dollar_2000')
    data = []
            # cursor.execute("drop table if exists %s" % mdb.escape_string(table_name))
            # query = 
            # INSERT INTO backup_db.date_manu_update SELECT * from logi.date_manu_update;"

        if cursor.execute("CREATE TABLE IF NOT EXISTS %s SELECT * FROM %s" % mdb.escape_string(table_name)):
                
            conn.commit()
            query =  "DROP TABLE IF EXISTS logi.%s;"
            cursor.execute(query,cell.value)
            conn.commit()
        
    PATH = 'Details/details.xlsx'
    if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
        wbb = openpyxl.load_workbook(PATH)
        sheetdetails = wbb["Sheet"]

        sheetdetails.append(rowss)
    else:
        wbb = openpyxl.Workbook()
        wss = wbb.active
        column_names = ["Table Names","line","path"]
        wss.append(column_names)
        sheetdetails = wbb["Sheet"]

        sheetdetails.append(rowss)
        wbb.save(PATH)
        
create_row()

# # switch = 0
# column_names = ["Deleted Tables"]
# def generate_report():
#     data = create_row()

#     PATH = './tmp/log_file.xlsx'
#     if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
#         wb = openpyxl.load_workbook('tmp/log_file.xlsx')
#         sheet = wb["Sheet"]

#         for row in data:
#             sheet.append(row)
#     else:
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.append(column_names)

#         sheet = wb["Sheet"]

#         for row in data:
#             sheet.append(row)
    
#     wb.save('tmp/log_file.xlsx')
# # schedule.every().hour.do(generate_report)
# # #schedule.every(1).minutes.do(generate_report)
# # while True:
# #     schedule.run_pending()
# #     time.sleep(1)


# generate_report()
