import time
import schedule
import openpyxl
import os
from mysql.connector import connect


conn = connect(host="127.0.0.1",user="root",passwd="12345678")
cursor = conn.cursor()
book = openpyxl.load_workbook('TableName/tablename.xlsx')
sheet = book.active
max_rows = sheet. max_row
def create_row():
    

    

    j=0
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=max_rows, max_col=1):
        i = 0
        print(j)
        j+=1
        for cell in row:
            if cell.value is None:
                continue
            print(cell.value, end=" ")
            print()
              
            newtablename ='backup_db.' + cell.value
            oldtablename ='logi.' + cell.value
            
            data = []
            # cursor.execute("drop table if exists %s" % mdb.escape_string(table_name))
            query = "CREATE TABLE IF NOT EXISTS " + newtablename + " SELECT * FROM " + oldtablename + ";"
            # INSERT INTO backup_db.date_manu_update SELECT * from logi.date_manu_update;"
            cursor.execute(query)
                
            conn.commit()
            query =  "DROP TABLE " + oldtablename + ";"
            cursor.execute(query)
            conn.commit()
            rowss = cell.value,1
            
            PATH = 'Details/details.xlsx'
            if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
                wbb = openpyxl.load_workbook(PATH)
                sheetdetails = wbb["Sheet"]

                sheetdetails.append(rowss)
            else:
                wbb = openpyxl.Workbook()
                wss = wbb.active
                column_names = ["Table Names","line","path"]
                wss.append(column_names)
                sheetdetails = wbb["Sheet"]

                sheetdetails.append(rowss)
                wbb.save(PATH)
            
create_row()

# # switch = 0
# column_names = ["Deleted Tables"]
# def generate_report():
#     data = create_row()

#     PATH = './tmp/log_file.xlsx'
#     if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
#         wb = openpyxl.load_workbook('tmp/log_file.xlsx')
#         sheet = wb["Sheet"]

#         for row in data:
#             sheet.append(row)
#     else:
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.append(column_names)

#         sheet = wb["Sheet"]

#         for row in data:
#             sheet.append(row)
    
#     wb.save('tmp/log_file.xlsx')
# # schedule.every().hour.do(generate_report)
# # #schedule.every(1).minutes.do(generate_report)
# # while True:
# #     schedule.run_pending()
# #     time.sleep(1)


# generate_report()
